#!/usr/bin/env python3
import os
import sys
import json
import requests
import pandas as pd
from datetime import datetime
from akamai.edgegrid import EdgeGridAuth, EdgeRc
from cryptography import x509
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import time  # For retry wait
import subprocess  # For fetching full user name on macOS

# ---------------- CONFIG ----------------
SECTION = "default"
INPUT_SHEET_NAME = "CSR Generator"
SUMMARY_FILE_NAME = "csr_summary.xlsx"

VERBOSE = True  # CSR PEM will NOT be printed

ENROLLMENTS_HEADERS = {
    "accept": "application/vnd.akamai.cps.enrollments.v11+json"
}

CSR_HEADERS = {
    "accept": "application/vnd.akamai.cps.csr.v2+json"
}

# ---------------- PATH HANDLING ----------------
def get_app_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

APP_DIR = get_app_dir()

# ---------------- GREETING ----------------
try:
    # macOS full user name
    user_name = subprocess.check_output(["id", "-F"], text=True).strip()
    if not user_name:
        user_name = "User"
except Exception:
    user_name = "User"

print(f"\n👋 Hi {user_name}\n")

# ---------------- EDGEGRID ----------------
edgerc_path = os.path.expanduser("~/.edgerc")
if not os.path.exists(edgerc_path):
    print(f"❌ Edgerc file not found at {edgerc_path}. Exiting...")
    sys.exit(1)

edgerc = EdgeRc(edgerc_path)
if not edgerc.has_section(SECTION):
    print(f"❌ Section '{SECTION}' not found in {edgerc_path}. Exiting...")
    sys.exit(1)

try:
    BASE_URL = f"https://{edgerc.get(SECTION, 'host')}"
    session = requests.Session()
    session.auth = EdgeGridAuth.from_edgerc(edgerc, SECTION)
except Exception as e:
    print(f"❌ Failed to read Edgerc section '{SECTION}': {e}")
    sys.exit(1)

# ---------------- HELPERS ----------------
def call_api(url, headers):
    return session.get(url, headers=headers)

def parse_csr(csr_pem):
    cert = x509.load_pem_x509_csr(csr_pem.encode())
    cn_attr = cert.subject.get_attributes_for_oid(x509.oid.NameOID.COMMON_NAME)
    cn = cn_attr[0].value if cn_attr else None

    org_attr = cert.subject.get_attributes_for_oid(x509.oid.NameOID.ORGANIZATION_NAME)
    org = org_attr[0].value if org_attr else None

    sans = []
    for ext in cert.extensions:
        if isinstance(ext.value, x509.SubjectAlternativeName):
            sans.extend(ext.value.get_values_for_type(x509.DNSName))

    return cn, org, sans

def save_csr_files(base_dir, key_type, cn, csr_pem):
    key_dir = os.path.join(base_dir, key_type)
    os.makedirs(key_dir, exist_ok=True)
    safe_cn = cn.replace("*", "_wildcard_") if cn else "unknown_cn"
    pem_path = os.path.join(key_dir, f"{safe_cn}_{key_type}.pem")
    txt_path = os.path.join(key_dir, f"{safe_cn}_{key_type}.txt")

    with open(pem_path, "w") as f:
        f.write(csr_pem)
    with open(txt_path, "w") as f:
        f.write(csr_pem)

def find_enrollment_by_slot(enrollments, slot_id):
    for e in enrollments:
        slots = (
            e.get("assignedSlots", []) +
            e.get("productionSlots", []) +
            e.get("stagingSlots", [])
        )
        if slot_id in slots:
            enrollment_id = e["location"].split("/")[-1]
            pending = e.get("pendingChanges", [])
            change_id = pending[0]["location"].split("/")[-1] if pending else None
            return enrollment_id, change_id
    return None, None

# ---------------- MAIN ----------------
def main():
    print("\n 🚀 Welcome to CPS Excel-Driven CSR Retrieval & Validation Tool 🚀\n\n")

    input_excel = os.path.join(APP_DIR, "csr_generator.xlsx")
    if not os.path.exists(input_excel):
        print(f"❌ Base sheet not found at {input_excel}")
        return

    df = pd.read_excel(input_excel, sheet_name=INPUT_SHEET_NAME)
    print(f"✅ Loaded base sheet: {input_excel}")

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    base_output_dir = os.path.join(APP_DIR, f"csr_run_{timestamp}")
    os.makedirs(base_output_dir, exist_ok=True)
    summary_file_name = f"csr_summary_{timestamp}.xlsx"

    summary_rows = []
    accounts_cache = {}
    processed_rows = 0

    for idx, row in df.iterrows():
        account_key = str(row.get("accountSwitchKey")).strip()

        # ✅ Handle missing CN cleanly
        input_cn = row.get("Common Name (CN)")
        if pd.isna(input_cn) or not str(input_cn).strip():
            input_cn = "unknown_cn"
        else:
            input_cn = str(input_cn).strip()

        slot_id = row.get("Slot")

        print("\n" + "=" * 80)
        print(f"Row {idx + 1} | CN: {input_cn}")

        if account_key in accounts_cache:
            enrollments = accounts_cache[account_key]
        else:
            url = f"{BASE_URL}/cps/v2/enrollments?accountSwitchKey={account_key}"
            resp = call_api(url, ENROLLMENTS_HEADERS)
            if resp.status_code != 200:
                print("❌ Failed to fetch enrollments")
                continue
            enrollments = resp.json().get("enrollments", [])
            accounts_cache[account_key] = enrollments

        enrollment_id = change_id = None

        if not pd.isna(slot_id):
            enrollment_id, change_id = find_enrollment_by_slot(enrollments, int(slot_id))
        else:
            for e in enrollments:
                if e.get("csr", {}).get("cn") == input_cn:
                    enrollment_id = e["location"].split("/")[-1]
                    pending = e.get("pendingChanges", [])
                    change_id = pending[0]["location"].split("/")[-1] if pending else None
                    break

        if not enrollment_id or not change_id:
            print("❌ Enrollment or Change ID not found")
            continue

        processed_rows += 1
        print(f"Enrollment ID: {enrollment_id}")
        print(f"Change ID: {change_id}")

        csr_url = (
            f"{BASE_URL}/cps/v2/enrollments/{enrollment_id}"
            f"/changes/{change_id}/input/info/third-party-csr"
            f"?accountSwitchKey={account_key}"
        )

        # ---------------- RETRY LOGIC ----------------
        max_retries = 3
        csrs = []
        for attempt in range(1, max_retries + 1):
            csr_resp = call_api(csr_url, CSR_HEADERS)
            if csr_resp.status_code == 200:
                csrs = csr_resp.json().get("csrs", [])
                if csrs:
                    break
            print(f"⚠️ CSR not ready yet, retrying ({attempt}/{max_retries}) ...")
            time.sleep(5)
        else:
            print("❌ CSR fetch failed after retries")
            continue

        for csr_entry in csrs:
            key_type = csr_entry["keyAlgorithm"]
            csr_pem = csr_entry["csr"].replace("\\n", "\n")
            try:
                csr_cn, org, sans = parse_csr(csr_pem)
                result = "PASS"
            except Exception:
                csr_cn = org = None
                sans = []
                result = "FAIL"

            print(f"\n🔐 {key_type} CSR")
            print(f"CSR Common Name (CN): {csr_cn}")
            print(f"SAN Present: {'Yes' if sans else 'No'}")
            print(f"SAN Count: {len(sans)}")
            if sans:
                print("SANs:")
                for s in sans:
                    print(f" - {s}")

            save_csr_files(base_output_dir, key_type, csr_cn or input_cn, csr_pem)

            summary_rows.append({
                "AccountSwitchKey": account_key,
                "SlotId": slot_id,
                "inputCommonName": input_cn,
                "CSR_CommonName": csr_cn,
                "Organization": org,
                "SAN_Count": len(sans),
                "SAN_List": ", ".join(sans),
                "keyType": key_type,
                "CSR_Result": result,
                "EnrollmentId": enrollment_id,
                "ChangeId": change_id
            })

    # ---------------- SAVE SUMMARY EXCEL WITH STYLING ----------------
    summary_path = os.path.join(base_output_dir, summary_file_name)
    df_summary = pd.DataFrame(summary_rows)
    df_summary.to_excel(summary_path, index=False)

    wb = load_workbook(summary_path)
    ws = wb.active

    header_fill = PatternFill(start_color="B0E0E6", end_color="B0E0E6", fill_type="solid")
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_cells in ws.columns:
        col_letter = column_cells[0].column_letter
        if column_cells[0].value == "SAN_List":
            ws.column_dimensions[col_letter].width = 40
        else:
            max_length = 0
            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    wb.save(summary_path)

    print("\n✅ CSR generation completed")
    print(f"📊 Excel rows processed: {processed_rows}")
    print(f"📄 Summary file: {summary_path}")
    print(f"📁 Output folder: {base_output_dir}")
    print("\n\n" + "-" * 60)
    print("Powered by Edge Squad | Internal Use Only")
    print("-" * 60 + "\n\n\n")

# ---------------- RUN ----------------
if __name__ == "__main__":
    main()
